import json
import urllib.request
import urllib.error
import openpyxl
from openpyxl.utils import get_column_letter
import os
import datetime

def get_hex_color(fill):
    if not fill or fill.fill_type != "solid" or not fill.start_color:
        return None
    rgb = getattr(fill.start_color, "rgb", None)
    if isinstance(rgb, str):
        rgb = rgb.strip()
        if len(rgb) == 8:
            return '#' + rgb[2:].lower()
        elif len(rgb) == 6:
            return '#' + rgb.lower()
    return None

def get_font_color(font):
    if not font or not font.color:
        return None
    rgb = getattr(font.color, "rgb", None)
    if isinstance(rgb, str):
        rgb = rgb.strip()
        if len(rgb) == 8:
            return '#' + rgb[2:].lower()
        elif len(rgb) == 6:
            return '#' + rgb.lower()
    return None

def sync_sheets():
    # Load config.json
    config_path = "config.json"
    if not os.path.exists(config_path):
        print("Error: config.json not found. Please create it with google_web_app_url.")
        return
        
    with open(config_path, 'r') as f:
        config = json.load(f)
        
    url = config.get("google_web_app_url")
    if not url or "YOUR_URL" in url:
        print("Error: Please set google_web_app_url in config.json.")
        return
        
    # Auto-sync config.js for the browser dashboard app (bypasses browser file:// CORS restrictions)
    config_js_path = "config.js"
    needs_write = True
    if os.path.exists(config_js_path):
        try:
            with open(config_js_path, 'r') as f:
                js_content = f.read()
                if url in js_content:
                    needs_write = False
        except:
            pass
    if needs_write:
        try:
            with open(config_js_path, 'w') as f:
                f.write(f'window.APP_CONFIG = {{\n  "google_web_app_url": "{url}"\n}};\n')
        except Exception as e:
            print(f"Warning: Could not write config.js: {e}")
        
    file_name = "Janvi_Consolidated_Orders_2026_2027.xlsx"
    if not os.path.exists(file_name):
        print(f"Error: Local Excel workbook '{file_name}' not found. Run generate_sheet.py first.")
        return
        
    print("Reading local Excel workbook (extracting values, merges, and layout)...")
    wb = openpyxl.load_workbook(file_name, data_only=False)
    
    payload = {"sheets": {}}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        grid = []
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        for r in range(1, max_row + 1):
            row_data = []
            for c in range(1, max_col + 1):
                cell = sheet.cell(row=r, column=c)
                val = cell.value
                
                if val is None:
                    val = ""
                elif isinstance(val, (datetime.datetime, datetime.date)):
                    val = val.strftime('%Y-%m-%d')
                elif isinstance(val, (int, float, str)):
                    pass
                else:
                    val = str(val)
                    
                # Extract style properties
                bg_color = get_hex_color(cell.fill)
                font_color = get_font_color(cell.font)
                is_bold = bool(cell.font.bold) if cell.font else False
                
                align = cell.alignment.horizontal if cell.alignment else "left"
                if align not in ["left", "center", "right"]:
                    align = "left"
                    
                is_wrapped = bool(cell.alignment.wrap_text) if cell.alignment else False
                
                # Lowercase number format for Google Sheets compatibility (e.g. YYYY-MM-DD -> yyyy-mm-dd)
                num_fmt = cell.number_format.lower() if cell.number_format else "@"
                
                row_data.append({
                    "value": val,
                    "bg": bg_color,
                    "color": font_color,
                    "bold": is_bold,
                    "align": align,
                    "num_format": num_fmt,
                    "wrap": is_wrapped
                })
            grid.append(row_data)
            
        # Get merged cell ranges
        merged_ranges = [str(r) for r in sheet.merged_cells.ranges]
        
        # Get column widths
        column_widths = {}
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            width = sheet.column_dimensions[col_letter].width
            if width is None:
                width = 12
            column_widths[str(col_idx)] = width
            
        # Get row heights
        row_heights = {}
        for r in range(1, max_row + 1):
            height = sheet.row_dimensions[r].height
            if height is not None:
                row_heights[str(r)] = height
        
        payload["sheets"][sheet_name] = {
            "grid": grid,
            "merged_ranges": merged_ranges,
            "column_widths": column_widths,
            "row_heights": row_heights
        }
        
    print("Uploading data, styles, and layout to Google Sheets web app...")
    try:
        data_bytes = json.dumps(payload).encode('utf-8')
        req = urllib.request.Request(
            url, 
            data=data_bytes, 
            headers={'Content-Type': 'application/json'}
        )
        with urllib.request.urlopen(req) as response:
            res_body = response.read().decode('utf-8')
            res_json = json.loads(res_body)
            if res_json.get("status") == "success":
                print("Successfully synced all sheets, styles, and layouts to Google Sheets!")
            else:
                print(f"Sync failed: {res_json.get('message')}")
    except urllib.error.HTTPError as e:
        print(f"HTTP Error: {e.code} - {e.reason}")
        try:
            err_body = e.read().decode('utf-8')
            print(f"Details: {err_body}")
        except:
            pass
    except urllib.error.URLError as e:
        print(f"Connection Error: {e.reason}")
    except Exception as e:
        print(f"Error during sync: {e}")

if __name__ == "__main__":
    sync_sheets()
