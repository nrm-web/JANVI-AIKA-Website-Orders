import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import datetime
import random
import re

def generate_synthetic_4see_sheet():
    output_filename = "Janvi_Consolidated_Orders_Synthetic_4see.xlsx"
    print(f"Generating synthetic orders workbook for 4see purpose: {output_filename}...")
    
    # ----------------------------------------------------
    # SECTION 1: SYNTHETIC DATA GENERATION
    # ----------------------------------------------------
    random.seed(42)  # For reproducible synthetic data
    
    first_names = [
        "Ananya", "Priya", "Meera", "Sanjana", "Kavya", "Deepika", "Ritu", "Neha", 
        "Shruti", "Pooja", "Aishwarya", "Divya", "Radhika", "Tanvi", "Swati", 
        "Aditi", "Anushka", "Isha", "Niharika", "Sneha", "Simran", "Tara", "Vaidehi"
    ]
    last_names = [
        "Sharma", "Patel", "Reddy", "Gupta", "Iyer", "Joshi", "Verma", "Kapoor",
        "Nair", "Rao", "Sen", "Menon", "Choudhary", "Bhatia", "Kulkarni", "Deshmukh",
        "Agarwal", "Mehta", "Singh", "Banerjee", "Hegde", "Pillai", "Shah"
    ]
    
    products = [
        ("Pure Silk Anarkali Suit Set", "A", "ANARKALI", 4999.0),
        ("Embroidered Georgette Lehenga Choli", "L", "LEHENGA", 7499.0),
        ("Designer Partywear Sharara Set", "SHA", "SHARARA", 3999.0),
        ("Printed Cotton Chudidhar Set", "C", "CHUDIDHAR", 1999.0),
        ("Floral Organza Long Gown", "LG", "LONG GOWN", 4499.0),
        ("Velvet Festive Co-Ord Set", "CORD", "CO-ORD SET", 3299.0),
        ("Traditional Silk Half Saree Lehenga", "HSL", "HALF SAREE LEHENGA", 8999.0),
        ("Casual Printed Top & Bottom", "TOP", "TOPS", 1299.0),
        ("Chanderi Kurti Chudidhar Set", "C", "CHUDIDHAR", 2499.0),
        ("Sequined Bridal Lehenga Choli", "L", "LEHENGA", 9999.0)
    ]
    
    cities_pins = [
        ("Bengaluru", "560001"), ("Mumbai", "400001"), ("Hyderabad", "500001"),
        ("Chennai", "600001"), ("Delhi", "110001"), ("Pune", "411001"),
        ("Ahmedabad", "380001"), ("Jaipur", "302001"), ("Surat", "395001"),
        ("Kochi", "682001"), ("Kolkata", "700001"), ("Lucknow", "226001"),
        ("Coimbatore", "641001"), ("Chandigarh", "160001"), ("Indore", "452001")
    ]
    
    months_range = [
        (2026, 6, "Jun 2026", 18),
        (2026, 7, "Jul 2026", 20),
        (2026, 8, "Aug 2026", 22),
        (2026, 9, "Sep 2026", 19),
        (2026, 10, "Oct 2026", 25),
        (2026, 11, "Nov 2026", 24),
        (2026, 12, "Dec 2026", 21),
        (2027, 1, "Jan 2027", 17),
        (2027, 2, "Feb 2027", 18),
        (2027, 3, "Mar 2027", 16)
    ]
    
    synthetic_records = []
    order_counter = 2001
    
    for year, month, tab_name, count in months_range:
        for i in range(count):
            order_no = f"#{order_counter}"
            order_counter += 1
            
            cust_name = f"{random.choice(first_names)} {random.choice(last_names)}"
            item_title, prefix, category, base_price = random.choice(products)
            sku = f"{prefix}-{random.randint(100, 999)}"
            price = round(base_price * random.uniform(0.9, 1.1), 2)
            
            day = random.randint(1, 28)
            date_str = f"{year}-{month:02d}-{day:02d}"
            
            city, pin = random.choice(cities_pins)
            
            # Payment distribution
            pay_rnd = random.random()
            if pay_rnd < 0.60:
                pay_method = "Prepaid (Razorpay)"
                is_cod = False
            elif pay_rnd < 0.85:
                pay_method = "COD"
                is_cod = True
            else:
                pay_method = "Partial COD"
                is_cod = True
                
            prepaid_str = "No" if is_cod else "Yes"
            cod_str = "Yes" if is_cod else "No"
            
            # Status distribution
            stat_rnd = random.random()
            returned = False
            cod_denied = "No"
            comments = "-"
            
            if stat_rnd < 0.75:
                fulfillment_status = "DELIVERED"
            elif stat_rnd < 0.83:
                fulfillment_status = "RTO DELIVERED"
                returned = True
                if is_cod:
                    cod_denied = "Yes"
                    comments = "Customer Refused To Accept / Undelivered"
                else:
                    comments = "Address Not Found / RTO"
            elif stat_rnd < 0.88:
                fulfillment_status = "CANCELED"
                comments = "Order Canceled by Customer"
            elif stat_rnd < 0.94:
                fulfillment_status = "IN TRANSIT"
                comments = "En-route to Destination Hub"
            elif stat_rnd < 0.97:
                fulfillment_status = "OUT FOR DELIVERY"
                comments = "Out for Delivery with Courier"
            else:
                fulfillment_status = "NEW ORDER"
                comments = "Manifested & Awaiting Pickup"
                
            feedback_sent = "Yes" if fulfillment_status == "DELIVERED" else "No"
            feedback_rec = "Yes" if (fulfillment_status == "DELIVERED" and random.random() > 0.6) else "No"
            
            awb = f"1411{random.randint(10000000, 99999999)}"
            tracking_link = f"https://shiprocket.co/tracking/{awb}"
            
            synthetic_records.append({
                "Order No": order_no,
                "Customer Name": cust_name,
                "Items Ordered": item_title,
                "Date of Order": date_str,
                "Total Price": price,
                "Payment Method": pay_method,
                "Prepaid (Yes/No)": prepaid_str,
                "COD (Yes/No)": cod_str,
                "Returned (True/False)": returned,
                "COD Denies (Yes/No)": cod_denied,
                "Shiprocket Comments": comments,
                "City": city,
                "PIN Code": pin,
                "Fulfillment Status": fulfillment_status,
                "Feedback Link Sent (Yes/No)": feedback_sent,
                "Feedback Received (Yes/No)": feedback_rec,
                "AWB Code": awb,
                "Tracking Link": tracking_link,
                "SKU": sku,
                "Category": category
            })
            
    df_synthetic = pd.DataFrame(synthetic_records)
    
    # Sort chronologically descending
    df_synthetic['order_num_int'] = df_synthetic['Order No'].str.extract(r'(\d+)').astype(float).fillna(0).astype(int)
    df_synthetic = df_synthetic.sort_values(by=['Date of Order', 'order_num_int'], ascending=[False, False])
    df_synthetic = df_synthetic.drop(columns=['order_num_int'])
    
    print(f"Total synthetic 4see orders generated: {len(df_synthetic)}")
    
    # ----------------------------------------------------
    # SECTION 2: EXCEL WRITING & STYLING SETUP
    # ----------------------------------------------------
    writer = pd.ExcelWriter(output_filename, engine='openpyxl')
    
    font_family = "Segoe UI"
    color_primary = "1F4E78"    # Steel Blue
    color_border = "D3D3D3"     # Light Gray
    color_zebra = "F9FBFD"      # Alternate row fill
    
    header_fill = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    zebra_fill = PatternFill(start_color=color_zebra, end_color=color_zebra, fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    font_regular = Font(name=font_family, size=10)
    font_bold = Font(name=font_family, size=10, bold=True)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_side = Side(border_style="thin", color=color_border)
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    border_summary = Border(
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='double', color=color_primary)
    )
    
    def style_table(sheet, df):
        sheet.views.sheetView[0].showGridLines = True
        sheet.row_dimensions[1].height = 28
        
        for col_idx in range(1, len(df.columns) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border_all
            
        for row_idx in range(2, len(df) + 2):
            sheet.row_dimensions[row_idx].height = 20
            row_fill = zebra_fill if row_idx % 2 == 1 else white_fill
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.fill = row_fill
                cell.font = font_regular
                cell.border = border_all
                
                col_name = df.columns[col_idx - 1]
                
                if col_name in ["Order No", "Date of Order", "Payment Method", "Prepaid (Yes/No)", "COD (Yes/No)", "Returned (True/False)", "COD Denies (Yes/No)", "PIN Code", "Fulfillment Status", "Feedback Link Sent (Yes/No)", "Feedback Received (Yes/No)", "AWB Code", "Tracking Link", "SKU", "Category"]:
                    cell.alignment = align_center
                elif col_name == "Items Ordered":
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = align_left
                    
                if col_name == "Total Price":
                    cell.alignment = align_right
                    cell.number_format = '₹#,##0.00'
                elif col_name == "Date of Order":
                    cell.number_format = 'dd-mm-yyyy'
                elif col_name in ["PIN Code"]:
                    cell.number_format = '@'
                elif col_name == "Returned (True/False)":
                    cell.value = str(cell.value)
                    
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = cell.value
                val_str = str(val or '')
                if isinstance(val, float) and "Price" in df.columns[cell.column - 1]:
                    val_str = f"₹{val:,.2f}"
                max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # ----------------------------------------------------
    # SECTION 3: WRITE MASTER SHEET & MONTHLY SHEETS
    # ----------------------------------------------------
    df_master_output = df_synthetic.copy()
    df_master_output['Date of Order'] = pd.to_datetime(df_master_output['Date of Order']).dt.date
    df_master_output.to_excel(writer, sheet_name="Master Sheet", index=False)
    sheet_master = writer.sheets["Master Sheet"]
    style_table(sheet_master, df_master_output)
    
    last_row_master = len(df_master_output) + 1
    sum_row_master = last_row_master + 1
    sheet_master.row_dimensions[sum_row_master].height = 22
    sheet_master.cell(row=sum_row_master, column=1, value="Total Orders")
    c_count = sheet_master.cell(row=sum_row_master, column=2, value=f"=COUNTA(A2:A{last_row_master})")
    c_count.alignment = align_center
    
    sheet_master.cell(row=sum_row_master, column=4, value="Total Revenue").alignment = align_right
    c_rev = sheet_master.cell(row=sum_row_master, column=5, value=f"=SUM(E2:E{last_row_master})")
    c_rev.alignment = align_right
    c_rev.number_format = '₹#,##0.00'
    
    for col_idx in range(1, len(df_master_output.columns) + 1):
        cell = sheet_master.cell(row=sum_row_master, column=col_idx)
        cell.border = border_summary
        cell.font = font_bold
        
    for year, month, tab_name, _ in months_range:
        df_month = df_synthetic[
            (pd.to_datetime(df_synthetic["Date of Order"]).dt.year == year) & 
            (pd.to_datetime(df_synthetic["Date of Order"]).dt.month == month)
        ].copy()
        
        is_empty = df_month.empty
        if is_empty:
            df_month = pd.DataFrame(columns=df_synthetic.columns)
        else:
            df_month['Date of Order'] = pd.to_datetime(df_month['Date of Order']).dt.date
            
        df_month.to_excel(writer, sheet_name=tab_name, index=False)
        sheet = writer.sheets[tab_name]
        style_table(sheet, df_month)
        
        if not is_empty:
            last_row = len(df_month) + 1
            sum_row = last_row + 1
            sheet.row_dimensions[sum_row].height = 22
            
            sheet.cell(row=sum_row, column=1, value="Total Orders")
            c_count = sheet.cell(row=sum_row, column=2, value=f"=COUNTA(A2:A{last_row})")
            c_count.alignment = align_center
            
            sheet.cell(row=sum_row, column=4, value="Total Revenue").alignment = align_right
            c_rev = sheet.cell(row=sum_row, column=5, value=f"=SUM(E2:E{last_row})")
            c_rev.alignment = align_right
            c_rev.number_format = '₹#,##0.00'
            
            for col_idx in range(1, len(df_month.columns) + 1):
                cell = sheet.cell(row=sum_row, column=col_idx)
                cell.border = border_summary
                cell.font = font_bold

    # ----------------------------------------------------
    # SECTION 4: CREATE & STYLE DASHBOARD SHEET
    # ----------------------------------------------------
    workbook = writer.book
    dash_sheet = workbook.create_sheet("Dashboard", 0)
    dash_sheet.views.sheetView[0].showGridLines = True
    
    # Title Block
    dash_sheet.merge_cells("A1:H1")
    for col_idx in range(1, 9):
        cell = dash_sheet.cell(row=1, column=col_idx)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        
    title_cell = dash_sheet["A1"]
    title_cell.value = "JANVI AIKA Synthetic 4See Forecasting Orders Dashboard (June 2026 - March 2027)"
    title_cell.font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    dash_sheet.row_dimensions[1].height = 40
    
    # Calculations for Dashboard
    total_orders = len(df_synthetic)
    total_revenue = df_synthetic["Total Price"].sum()
    
    canceled_df = df_synthetic[df_synthetic["Fulfillment Status"].str.upper().str.contains("CANCELED|CANCELLED", na=False)]
    canceled_count = len(canceled_df)
    canceled_amount = canceled_df["Total Price"].sum()
    
    denied_mask = ((df_synthetic["Fulfillment Status"].str.upper().str.contains("RTO|DENIED", na=False)) | (df_synthetic["COD Denies (Yes/No)"] == "Yes")) & (~df_synthetic["Fulfillment Status"].str.upper().str.contains("CANCELED|CANCELLED", na=False))
    denied_df = df_synthetic[denied_mask]
    denied_count = len(denied_df)
    denied_amount = denied_df["Total Price"].sum()
    
    returned_mask = (df_synthetic["Returned (True/False)"] == True) & (~denied_mask) & (~df_synthetic["Fulfillment Status"].str.upper().str.contains("CANCELED|CANCELLED", na=False))
    returned_df = df_synthetic[returned_mask]
    returned_count = len(returned_df)
    returned_amount = returned_df["Total Price"].sum()
    
    net_revenue_val = total_revenue - denied_amount - returned_amount - canceled_amount
    net_revenue_count = total_orders - denied_count - returned_count - canceled_count
    
    successful_mask = (df_synthetic["Fulfillment Status"].str.upper().str.strip().isin(["DELIVERED", "SELF FULFILED", "FULFILLED"]))
    successful_mask = successful_mask & (~df_synthetic["Returned (True/False)"] == True) & (~denied_mask) & (~df_synthetic["Fulfillment Status"].str.upper().str.contains("CANCELED|CANCELLED", na=False))
    successful_df = df_synthetic[successful_mask]
    successful_count = len(successful_df)
    aov = (net_revenue_val / successful_count) if successful_count > 0 else 0
    
    cod_successful_df = successful_df[successful_df["COD (Yes/No)"] == "Yes"]
    cod_successful_count = len(cod_successful_df)
    cod_net_revenue = cod_successful_df["Total Price"].sum()
    cod_aov = (cod_net_revenue / cod_successful_count) if cod_successful_count > 0 else 0
    
    prepaid_successful_df = successful_df[successful_df["Prepaid (Yes/No)"] == "Yes"]
    prepaid_successful_count = len(prepaid_successful_df)
    prepaid_net_revenue = prepaid_successful_df["Total Price"].sum()
    prepaid_aov = (prepaid_net_revenue / prepaid_successful_count) if prepaid_successful_count > 0 else 0
    
    in_progress_statuses = ['IN TRANSIT', 'IN TRANSIT-EN-ROUTE', 'SHIPPED', 'PICKED UP', 'REACHED DESTINATION HUB', 'OUT FOR DELIVERY', 'UNDELIVERED-1ST ATTEMPT', 'IN TRANSIT-AT DESTINATION HUB', 'PICKUP SCHEDULED', 'NEW ORDER']
    in_transit_statuses = ['IN TRANSIT', 'IN TRANSIT-EN-ROUTE', 'SHIPPED', 'PICKED UP', 'REACHED DESTINATION HUB', 'OUT FOR DELIVERY', 'UNDELIVERED-1ST ATTEMPT', 'IN TRANSIT-AT DESTINATION HUB']
    pickup_statuses = ['PICKUP SCHEDULED', 'NEW ORDER']
    
    in_progress_df = df_synthetic[df_synthetic["Fulfillment Status"].str.upper().str.strip().isin(in_progress_statuses)]
    in_progress_count = len(in_progress_df)
    in_progress_amount = in_progress_df["Total Price"].sum()
    
    in_transit_df = df_synthetic[df_synthetic["Fulfillment Status"].str.upper().str.strip().isin(in_transit_statuses)]
    in_transit_count = len(in_transit_df)
    in_transit_amount = in_transit_df["Total Price"].sum()
    
    pickup_df = df_synthetic[df_synthetic["Fulfillment Status"].str.upper().str.strip().isin(pickup_statuses)]
    pickup_count = len(pickup_df)
    pickup_amount = pickup_df["Total Price"].sum()
    
    unfulfilled_df = df_synthetic[df_synthetic["Fulfillment Status"].str.upper().str.strip() == 'UNFULFILLED']
    unfulfilled_count = len(unfulfilled_df)
    unfulfilled_amount = unfulfilled_df["Total Price"].sum()
    
    card_fill = PatternFill(start_color="F9F9FB", end_color="F9F9FB", fill_type="solid")
    
    # ROW 1 OF KPIs: Primary Overview
    dash_sheet["B3"] = "TOTAL ORDERS"
    dash_sheet["B3"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["B3"].alignment = align_center
    dash_sheet["B3"].fill = card_fill
    dash_sheet["B4"] = total_orders
    dash_sheet["B4"].font = Font(name=font_family, size=16, bold=True, color=color_primary)
    dash_sheet["B4"].alignment = align_center
    dash_sheet["B4"].number_format = '#,##0'
    dash_sheet["B4"].fill = card_fill
    
    dash_sheet["C3"] = "TOTAL REVENUE"
    dash_sheet["C3"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["C3"].alignment = align_center
    dash_sheet["C3"].fill = card_fill
    dash_sheet["C4"] = total_revenue
    dash_sheet["C4"].font = Font(name=font_family, size=16, bold=True, color="1E8449")
    dash_sheet["C4"].alignment = align_center
    dash_sheet["C4"].number_format = '₹#,##0.00'
    dash_sheet["C4"].fill = card_fill
    
    dash_sheet["D3"] = "NET REVENUE"
    dash_sheet["D3"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["D3"].alignment = align_center
    dash_sheet["D3"].fill = card_fill
    dash_sheet["D4"] = net_revenue_val
    dash_sheet["D4"].font = Font(name=font_family, size=16, bold=True, color="1E8449")
    dash_sheet["D4"].alignment = align_center
    dash_sheet["D4"].number_format = '₹#,##0.00'
    dash_sheet["D4"].fill = card_fill
    
    dash_sheet["E3"] = "AVERAGE ORDER VALUE (AOV)"
    dash_sheet["E3"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["E3"].alignment = align_center
    dash_sheet["E3"].fill = card_fill
    dash_sheet["E4"] = f"~ ₹{aov:,.2f}"
    dash_sheet["E4"].font = Font(name=font_family, size=16, bold=True, color="1F618D")
    dash_sheet["E4"].alignment = align_center
    dash_sheet["E4"].fill = card_fill
    
    # ROW 2 OF KPIs: Revenue Stream Breakdown
    dash_sheet["B6"] = "COD NET REVENUE"
    dash_sheet["B6"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["B6"].alignment = align_center
    dash_sheet["B6"].fill = card_fill
    dash_sheet["B7"] = cod_net_revenue
    dash_sheet["B7"].font = Font(name=font_family, size=16, bold=True, color="D35400")
    dash_sheet["B7"].alignment = align_center
    dash_sheet["B7"].number_format = '₹#,##0.00'
    dash_sheet["B7"].fill = card_fill
    
    dash_sheet["C6"] = "AVG COD ORDER VALUE"
    dash_sheet["C6"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["C6"].alignment = align_center
    dash_sheet["C6"].fill = card_fill
    dash_sheet["C7"] = f"~ ₹{cod_aov:,.2f}"
    dash_sheet["C7"].font = Font(name=font_family, size=16, bold=True, color="D35400")
    dash_sheet["C7"].alignment = align_center
    dash_sheet["C7"].fill = card_fill
    
    dash_sheet["D6"] = "PREPAID NET REVENUE"
    dash_sheet["D6"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["D6"].alignment = align_center
    dash_sheet["D6"].fill = card_fill
    dash_sheet["D7"] = prepaid_net_revenue
    dash_sheet["D7"].font = Font(name=font_family, size=16, bold=True, color="1E8449")
    dash_sheet["D7"].alignment = align_center
    dash_sheet["D7"].number_format = '₹#,##0.00'
    dash_sheet["D7"].fill = card_fill
    
    dash_sheet["E6"] = "AVG PREPAID ORDER VALUE"
    dash_sheet["E6"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["E6"].alignment = align_center
    dash_sheet["E6"].fill = card_fill
    dash_sheet["E7"] = f"~ ₹{prepaid_aov:,.2f}"
    dash_sheet["E7"].font = Font(name=font_family, size=16, bold=True, color="1E8449")
    dash_sheet["E7"].alignment = align_center
    dash_sheet["E7"].fill = card_fill

    # ROW 3 OF KPIs: Returns & Denials
    dash_sheet["B9"] = "SUCCESSFUL DELIVERED"
    dash_sheet["B9"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["B9"].alignment = align_center
    dash_sheet["B9"].fill = card_fill
    dash_sheet["B10"] = f"{successful_count} Orders"
    dash_sheet["B10"].font = Font(name=font_family, size=16, bold=True, color="1E8449")
    dash_sheet["B10"].alignment = align_center
    dash_sheet["B10"].fill = card_fill
    
    dash_sheet["C9"] = "DENIED ORDERS (COD/RTO)"
    dash_sheet["C9"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["C9"].alignment = align_center
    dash_sheet["C9"].fill = card_fill
    dash_sheet["C10"] = f"{denied_count} Orders (₹{denied_amount:,.2f})"
    dash_sheet["C10"].font = Font(name=font_family, size=14, bold=True, color="C0392B")
    dash_sheet["C10"].alignment = align_center
    dash_sheet["C10"].fill = card_fill

    dash_sheet["D9"] = "CUSTOMER RETURNED"
    dash_sheet["D9"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["D9"].alignment = align_center
    dash_sheet["D9"].fill = card_fill
    dash_sheet["D10"] = f"{returned_count} Orders (₹{returned_amount:,.2f})"
    dash_sheet["D10"].font = Font(name=font_family, size=14, bold=True, color="E67E22")
    dash_sheet["D10"].alignment = align_center
    dash_sheet["D10"].fill = card_fill

    dash_sheet["E9"] = "CANCELED ORDERS"
    dash_sheet["E9"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["E9"].alignment = align_center
    dash_sheet["E9"].fill = card_fill
    dash_sheet["E10"] = f"{canceled_count} Orders (₹{canceled_amount:,.2f})"
    dash_sheet["E10"].font = Font(name=font_family, size=14, bold=True, color="7F8C8D")
    dash_sheet["E10"].alignment = align_center
    dash_sheet["E10"].fill = card_fill

    # ROW 4 OF KPIs: Active Pipeline
    dash_sheet["B12"] = "ACTIVE SHIPPING PIPELINE"
    dash_sheet["B12"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["B12"].alignment = align_center
    dash_sheet["B12"].fill = card_fill
    dash_sheet["B13"] = f"{in_progress_count} Orders (₹{in_progress_amount:,.2f})"
    dash_sheet["B13"].font = Font(name=font_family, size=14, bold=True, color="2980B9")
    dash_sheet["B13"].alignment = align_center
    dash_sheet["B13"].fill = card_fill

    dash_sheet["C12"] = "IN TRANSIT / EN ROUTE"
    dash_sheet["C12"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["C12"].alignment = align_center
    dash_sheet["C12"].fill = card_fill
    dash_sheet["C13"] = f"{in_transit_count} Orders (₹{in_transit_amount:,.2f})"
    dash_sheet["C13"].font = Font(name=font_family, size=14, bold=True, color="2980B9")
    dash_sheet["C13"].alignment = align_center
    dash_sheet["C13"].fill = card_fill

    dash_sheet["D12"] = "PICKUP / MANIFESTED"
    dash_sheet["D12"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["D12"].alignment = align_center
    dash_sheet["D12"].fill = card_fill
    dash_sheet["D13"] = f"{pickup_count} Orders (₹{pickup_amount:,.2f})"
    dash_sheet["D13"].font = Font(name=font_family, size=14, bold=True, color="8E44AD")
    dash_sheet["D13"].alignment = align_center
    dash_sheet["D13"].fill = card_fill

    dash_sheet["E12"] = "UNFULFILLED ORDERS"
    dash_sheet["E12"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["E12"].alignment = align_center
    dash_sheet["E12"].fill = card_fill
    dash_sheet["E13"] = f"{unfulfilled_count} Orders (₹{unfulfilled_amount:,.2f})"
    dash_sheet["E13"].font = Font(name=font_family, size=14, bold=True, color="7F8C8D")
    dash_sheet["E13"].alignment = align_center
    dash_sheet["E13"].fill = card_fill

    # Style card borders
    card_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for r in [3, 4, 6, 7, 9, 10, 12, 13]:
        for c in [2, 3, 4, 5]:
            dash_sheet.cell(row=r, column=c).border = card_border

    # Monthly Performance Summary Table on Dashboard
    dash_sheet.merge_cells("A16:D16")
    summary_title = dash_sheet["A16"]
    summary_title.value = "MONTHLY PERFORMANCE BREAKDOWN"
    summary_title.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    summary_title.fill = header_fill
    summary_title.alignment = Alignment(horizontal="center", vertical="center")
    dash_sheet.row_dimensions[16].height = 24
    
    headers = ["Month", "Total Orders", "Total Revenue", "AOV"]
    for col_idx, text in enumerate(headers, 1):
        cell = dash_sheet.cell(row=17, column=col_idx, value=text)
        cell.font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_all
    dash_sheet.row_dimensions[17].height = 22
    
    row_pos = 18
    for year, month, tab_name, _ in months_range:
        df_m = df_synthetic[
            (pd.to_datetime(df_synthetic["Date of Order"]).dt.year == year) & 
            (pd.to_datetime(df_synthetic["Date of Order"]).dt.month == month)
        ]
        m_orders = len(df_m)
        m_rev = df_m["Total Price"].sum()
        m_aov = (m_rev / m_orders) if m_orders > 0 else 0
        
        m_fill = zebra_fill if (row_pos % 2 == 1) else white_fill
        
        c1 = dash_sheet.cell(row=row_pos, column=1, value=tab_name)
        c1.alignment = align_center
        c1.fill = m_fill
        c1.font = font_regular
        c1.border = border_all
        
        c2 = dash_sheet.cell(row=row_pos, column=2, value=m_orders)
        c2.alignment = align_center
        c2.fill = m_fill
        c2.font = font_regular
        c2.border = border_all
        c2.number_format = '#,##0'
        
        c3 = dash_sheet.cell(row=row_pos, column=3, value=m_rev)
        c3.alignment = align_right
        c3.fill = m_fill
        c3.font = font_regular
        c3.border = border_all
        c3.number_format = '₹#,##0.00'
        
        c4 = dash_sheet.cell(row=row_pos, column=4, value=m_aov)
        c4.alignment = align_right
        c4.fill = m_fill
        c4.font = font_regular
        c4.border = border_all
        c4.number_format = '₹#,##0.00'
        
        dash_sheet.row_dimensions[row_pos].height = 20
        row_pos += 1

    # Add summary row at bottom of Monthly Performance Breakdown
    dash_sheet.row_dimensions[row_pos].height = 22
    dash_sheet.cell(row=row_pos, column=1, value="Total").alignment = align_center
    dash_sheet.cell(row=row_pos, column=2, value=f"=SUM(B18:B{row_pos-1})").alignment = align_center
    c_tot_rev = dash_sheet.cell(row=row_pos, column=3, value=f"=SUM(C18:C{row_pos-1})")
    c_tot_rev.alignment = align_right
    c_tot_rev.number_format = '₹#,##0.00'
    
    c_avg_aov = dash_sheet.cell(row=row_pos, column=4, value=f"=C{row_pos}/B{row_pos}")
    c_avg_aov.alignment = align_right
    c_avg_aov.number_format = '₹#,##0.00'
    
    for c_i in range(1, 5):
        cell = dash_sheet.cell(row=row_pos, column=c_i)
        cell.font = font_bold
        cell.border = border_summary

    dash_sheet.column_dimensions["A"].width = 24
    dash_sheet.column_dimensions["B"].width = 28
    dash_sheet.column_dimensions["C"].width = 30
    dash_sheet.column_dimensions["D"].width = 30
    dash_sheet.column_dimensions["E"].width = 30
    dash_sheet.column_dimensions["F"].width = 16

    writer.close()
    print(f"Successfully generated synthetic Excel workbook: {output_filename}")

if __name__ == "__main__":
    generate_synthetic_4see_sheet()
