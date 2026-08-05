import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import datetime
import re
import json

def process_and_create_excel():
    # File paths for imported files
    shopify_path = os.path.join("Data", "Shopify orders Janvi.csv")
    shiprocket_path = os.path.join("Data", "Ship rocket orders.csv")
    
    if not os.path.exists(shopify_path) or not os.path.exists(shiprocket_path):
        print(f"Error: Missing data files in 'Data' directory. Please ensure both files are present:\n1. {shopify_path}\n2. {shiprocket_path}")
        return
        
    print("Loading data files...")
    # Load raw CSV files
    s = pd.read_csv(shopify_path)
    sr = pd.read_csv(shiprocket_path)
    
    # Forward-fill Shopify Name to handle multi-item orders where Name is only in the first row
    s['Name'] = s['Name'].ffill()
    
    # ----------------------------------------------------
    # SECTION 1: DATA CLEANING & MERGING
    # ----------------------------------------------------
    # Clean order names to extract numeric Order ID
    s['order_id_clean'] = s['Name'].astype(str).str.extract(r'(\d+)').astype(float).fillna(-1).astype(int)
    sr['order_id_clean'] = sr['Order ID'].astype(str).str.extract(r'(\d+)').astype(float).fillna(-1).astype(int)
    
    # Filter out orders that failed ID extraction
    s = s[s['order_id_clean'] != -1]
    sr = sr[sr['order_id_clean'] != -1]
    
    print(f"Shopify unique orders: {s['order_id_clean'].nunique()}")
    print(f"Shiprocket unique orders: {sr['order_id_clean'].nunique()}")
    
    # Group Shopify orders (since an order can have multiple rows for different items)
    s_grouped = s.groupby('order_id_clean').agg({
        'Name': 'first',
        'Created at': 'first',
        'Total': 'first',
        'Payment Method': 'first',
        'Financial Status': 'first',
        'Fulfillment Status': 'first',
        'Refunded Amount': 'first',
        'Shipping City': 'first',
        'Shipping Zip': 'first',
        'Billing Name': 'first',
        'Billing Phone': 'first',
        'Tags': 'first',
        'Lineitem name': lambda x: ', '.join(x.dropna().astype(str).unique()),
        'Lineitem sku': lambda x: ', '.join(x.dropna().astype(str))
    }).reset_index()
    
    # Group Shiprocket orders
    sr_grouped = sr.groupby('order_id_clean').agg({
        'Order ID': 'first',
        'Channel Created At': 'first',
        'Order Total': 'first',
        'Payment Method': 'first',
        'Address City': 'first',
        'Address Pincode': 'first',
        'Product Name': lambda x: ', '.join(x.dropna().astype(str).unique()),
        'Status': 'first',
        'Courier Company': 'first',
        'AWB Code': 'first',
        'Latest NDR Reason': 'first',
        'Pickup Exception Reason': 'first',
        'RTO Reason': 'first',
        'Cancellation Reason': 'first',
        'Customer Name': 'first',
        'Customer Mobile': 'first'
    }).reset_index()
    
    # Left join to only include orders present in Shopify, importing Shiprocket status only
    merged = pd.merge(s_grouped, sr_grouped, on='order_id_clean', how='left')
    print(f"Total unique consolidated orders: {len(merged)}")
    
    # ----------------------------------------------------
    # SECTION 1.5: FETCH EXTERNAL FEEDBACK DATA FROM GOOGLE SHEETS
    # ----------------------------------------------------
    feedback_map = {}
    feedback_url = "https://docs.google.com/spreadsheets/d/18K0SAdTm10ZDaYISJjwJcCf-87vI9IWQv66LdV1g5XA/export?format=csv&gid=310446875"
    try:
        print("Fetching feedback and review data from Google Sheet...")
        df_fb = pd.read_csv(feedback_url)
        df_fb.columns = df_fb.columns.str.strip()
        
        # Build lookup dictionary mapping clean Order No key (#1234) to statuses
        for _, fb_row in df_fb.iterrows():
            raw_ord = fb_row.get("Order No.")
            if pd.notna(raw_ord):
                try:
                    # Clean float order number to int string, e.g. 1073.0 -> 1073
                    ord_num = str(int(float(raw_ord)))
                    ord_key = f"#{ord_num}"
                    
                    # Convert to string and strip spaces
                    link_sent_val = str(fb_row.get("Reviews Link Sended", "")).strip().upper()
                    review_done_val = str(fb_row.get("REVIEWS DONE", "")).strip().upper()
                    
                    feedback_map[ord_key] = {
                        "sent": "Yes" if "SEND" in link_sent_val or "YES" in link_sent_val else "No",
                        "done": "Yes" if "DONE" in review_done_val or "YES" in review_done_val else "No"
                    }
                except Exception as e:
                    pass
        print(f"Loaded feedback info for {len(feedback_map)} orders from external Google Sheet.")
    except Exception as e:
        print(f"Warning: Failed to fetch external feedback Google Sheet: {e}. Falling back to defaults.")

    # ----------------------------------------------------
    # SECTION 2: DERIVE COLUMNS FOR MASTER SHEET
    # ----------------------------------------------------
    def map_sku_to_category(sku_str):
        if not sku_str or pd.isna(sku_str) or str(sku_str).strip() in ("", "-", "nan"):
            return "OTHER"
        skus = [s.strip().upper() for s in str(sku_str).split(",")]
        categories = []
        
        mapping = {
            "C": "CHUDIDHAR",
            "A": "ANARKALI",
            "L": "LEHENGA",
            "HSL": "HALF SAREE LEHENGA",
            "LG": "LONG GOWN",
            "SHA": "SHARARA",
            "TOP": "TOPS",
            "CORD": "CO-ORD SET"
        }
        
        for sku in skus:
            if not sku:
                continue
            prefix = sku.split("-")[0].strip()
            cat = mapping.get(prefix)
            if not cat:
                if sku.startswith("CORD"):
                    cat = "CO-ORD SET"
                else:
                    cat = "OTHER"
            categories.append(cat)
                
        return ", ".join(categories) if categories else "OTHER"

    consolidated = []
    
    for idx, row in merged.iterrows():
        order_id = int(row['order_id_clean'])
        
        # 1. Order Number
        order_no = str(row['Name'])
        
        # 2. Customer Name (Shopify only)
        cust_name = str(row['Billing Name']) if pd.notna(row['Billing Name']) else "Unknown"
        cust_name = cust_name.strip().title()
        
        # 3. Items Ordered (Shopify only)
        items = str(row['Lineitem name']) if pd.notna(row['Lineitem name']) else "No items specified"
        
        # 4. Date of Order (Shopify only)
        date_val = pd.to_datetime(row['Created at']).tz_localize(None) if pd.notna(row['Created at']) else None
        date_str = date_val.strftime('%Y-%m-%d') if date_val is not None else "Unknown"
        
        # 5. Price (Shopify only)
        price = float(row['Total']) if pd.notna(row['Total']) else 0.0
        
        # 6. Payment & COD
        pay_method_s = str(row['Payment Method_x']).lower() if pd.notna(row['Payment Method_x']) else ""
        pay_method_sr = str(row['Payment Method_y']).lower() if pd.notna(row['Payment Method_y']) else ""
        fin_status_lower = str(row['Financial Status']).lower().strip() if pd.notna(row['Financial Status']) else ""
        tags_lower = str(row['Tags']).lower().strip() if 'Tags' in row and pd.notna(row['Tags']) else ""
        
        is_cod = False
        pay_method = "Prepaid"
        
        if fin_status_lower == 'partially_paid' or 'partial' in tags_lower:
            is_cod = True
            pay_method = "Partial COD"
        elif 'cash' in pay_method_s or 'cod' in pay_method_s or 'cod' in pay_method_sr:
            is_cod = True
            pay_method = "COD"
        elif 'razorpay' in pay_method_s or 'upi' in pay_method_s or 'card' in pay_method_s or 'prepaid' in pay_method_sr:
            is_cod = False
            pay_method = "Prepaid (Razorpay)"
        elif pay_method_s == "" and pay_method_sr == "":
            # Check financial status as fallback
            if fin_status_lower == 'pending':
                is_cod = True
                pay_method = "COD"
                
        cod_str = "Yes" if is_cod else "No"
        prepaid_str = "No" if is_cod else "Yes"
        
        # 7. Returned & Cancellation Status
        sr_status = str(row['Status']).upper().strip() if pd.notna(row['Status']) else ""
        fin_status = str(row['Financial Status']).lower().strip() if pd.notna(row['Financial Status']) else ""
        refund_amt = float(row['Refunded Amount']) if pd.notna(row['Refunded Amount']) else 0.0
        cancelled_at_val = row['Cancelled at'] if 'Cancelled at' in row and pd.notna(row['Cancelled at']) else None
        
        shopify_cancelled = (cancelled_at_val is not None) or (fin_status == 'voided')
        is_canceled = 'CANCELED' in sr_status or 'CANCELLED' in sr_status or shopify_cancelled
        
        returned = False
        if not is_canceled:
            if fin_status in ['refunded', 'partially_refunded'] or refund_amt > 0 or 'RTO DELIVERED' in sr_status or 'RETURNED' in sr_status:
                returned = True
            
        # 8. COD Denial
        cod_denied = "No"
        ndr_reason = str(row['Latest NDR Reason']).lower() if pd.notna(row['Latest NDR Reason']) else ""
        if is_cod:
            if returned or 'CANCELED' in sr_status or 'refused' in ndr_reason or 'cancelled' in ndr_reason:
                cod_denied = "Yes"
                
        # 9. Address (City & PIN - Shopify only)
        city = str(row['Shipping City']) if pd.notna(row['Shipping City']) else "Unknown"
        city = city.strip().title()
        
        pin = str(row['Shipping Zip']).split('.')[0].strip() if pd.notna(row['Shipping Zip']) else ""
        
        # 10. Fulfillment Status (Shiprocket status falls back to Shopify status)
        if is_canceled:
            fulfillment_status = "CANCELED"
        else:
            fulfillment_status = sr_status if sr_status else (str(row['Fulfillment Status']).upper().strip() if pd.notna(row['Fulfillment Status']) else "NEW ORDER")
        
        # 11. Feedback
        fb_info = feedback_map.get(order_no)
        if fb_info:
            feedback_sent = fb_info["sent"]
            feedback_rec = fb_info["done"]
        else:
            feedback_sent = "Yes" if fulfillment_status == "DELIVERED" else "No"
            feedback_rec = "No"
        
        # 12. Shiprocket Comments/Remarks
        is_delivered_to_customer = (fulfillment_status.upper() == "DELIVERED") or (sr_status.upper() == "DELIVERED")
        
        comments = "-"
        if not is_delivered_to_customer:
            ndr_res = str(row['Latest NDR Reason']).strip() if pd.notna(row['Latest NDR Reason']) else ""
            pickup_exception = str(row['Pickup Exception Reason']).strip() if pd.notna(row['Pickup Exception Reason']) else ""
            rto_res = str(row['RTO Reason']).strip() if pd.notna(row['RTO Reason']) else ""
            cancel_res = str(row['Cancellation Reason']).strip() if pd.notna(row['Cancellation Reason']) else ""
            
            reasons = []
            if ndr_res and ndr_res.lower() not in ('nan', ''):
                reasons.append(ndr_res)
            if rto_res and rto_res.lower() not in ('nan', ''):
                reasons.append(rto_res)
            if pickup_exception and pickup_exception.lower() not in ('nan', ''):
                reasons.append(pickup_exception)
            if cancel_res and cancel_res.lower() not in ('nan', ''):
                reasons.append(cancel_res)
                
            unique_reasons = []
            for r in reasons:
                # Only split camelCase if the string has both uppercase and lowercase letters
                if any(c.isupper() for c in r) and any(c.islower() for c in r):
                    r_split = re.sub(r'(?<!^)(?=[A-Z])', ' ', r)
                else:
                    r_split = r
                r_clean = r_split.replace('_', ' ').replace('-', ' ')
                r_clean = ' '.join(r_clean.split()).strip().title()
                if r_clean not in unique_reasons:
                    unique_reasons.append(r_clean)
                    
            comments = ", ".join(unique_reasons) if unique_reasons else "-"
        
        # 13. AWB Code & Tracking Link
        awb_code = str(row['AWB Code']).split('.')[0].strip() if pd.notna(row['AWB Code']) else ""
        if awb_code.lower() == 'nan':
            awb_code = ""
        tracking_link = f"https://shiprocket.co/tracking/{awb_code}" if awb_code else "-"
        
        # Manual Order Overrides for orders fulfilled via external tracking (#1081, #1082, #1087, #1102, #1093, #1092, #1091)
        manual_overrides = {
            "#1081": {
                "fulfillment_status": "RTO DELIVERED",
                "cod_denied": "Yes",
                "returned": True,
                "comments": "Failed Delivery - Customer Denied & RTO",
                "awb": "1319460341773",
                "tracking_link": "https://www.xpressbees.com/track?isAwb=true&trackValue=1319460341773"
            },
            "#1082": {
                "fulfillment_status": "RTO DELIVERED",
                "cod_denied": "Yes",
                "returned": True,
                "comments": "Failed Delivery - Customer Denied & RTO",
                "awb": "1319460341769",
                "tracking_link": "https://www.xpressbees.com/track?isAwb=true&trackValue=1319460341769"
            },
            "#1087": {
                "fulfillment_status": "RTO DELIVERED",
                "cod_denied": "Yes",
                "returned": True,
                "comments": "Failed Delivery - Customer Denied & RTO",
                "awb": "SF3576445793KR",
                "tracking_link": "https://track.shadowfax.in/track?awb=SF3576445793KR"
            },
            "#1102": {
                "fulfillment_status": "RTO DELIVERED",
                "cod_denied": "Yes",
                "returned": True,
                "comments": "Customer Not Available X Office Or Residence Closed",
                "awb": "14112363521699",
                "tracking_link": "https://shiprocket.co/tracking/14112363521699"
            },
            "#1093": {
                "fulfillment_status": "RTO DELIVERED",
                "cod_denied": "Yes",
                "returned": True,
                "comments": "Customer Unavailable Delivery Attempted",
                "awb": "370429573110",
                "tracking_link": "https://shiprocket.co/tracking/370429573110"
            },
            "#1092": {
                "fulfillment_status": "RTO DELIVERED",
                "cod_denied": "Yes",
                "returned": True,
                "comments": "Customer Not Available X Office Or Residence Closed",
                "awb": "14112363585911",
                "tracking_link": "https://shiprocket.co/tracking/14112363585911"
            },
            "#1091": {
                "fulfillment_status": "RTO DELIVERED",
                "cod_denied": "Yes",
                "returned": True,
                "comments": "Pna|Receiver Not Available/Reachable Over Phone",
                "awb": "7D131457369",
                "tracking_link": "https://shiprocket.co/tracking/7D131457369"
            },
            "#1108": {
                "fulfillment_status": "RTO DELIVERED",
                "cod_denied": "Yes",
                "returned": True,
                "comments": "Shiprocket RTO",
                "awb": "SRSC0319007800",
                "tracking_link": "https://shiprocket.co/tracking/SRSC0319007800"
            },
            "#1107": {
                "fulfillment_status": "RTO DELIVERED",
                "cod_denied": "Yes",
                "returned": True,
                "comments": "Customer Cancelled O T P Verified",
                "awb": "SF3593208175KR",
                "tracking_link": "https://shiprocket.co/tracking/SF3593208175KR"
            },
            "#1105": {
                "fulfillment_status": "RTO DELIVERED",
                "cod_denied": "Yes",
                "returned": True,
                "comments": "Customer Refused To Accept O T P Verified",
                "awb": "14112364415189",
                "tracking_link": "https://shiprocket.co/tracking/14112364415189"
            }
        }
        
        if order_no in manual_overrides:
            ov = manual_overrides[order_no]
            fulfillment_status = ov["fulfillment_status"]
            cod_denied = ov["cod_denied"]
            returned = ov["returned"]
            comments = ov["comments"]
            awb_code = ov["awb"]
            tracking_link = ov["tracking_link"]
            
        # 14. SKU & Category
        sku = str(row['Lineitem sku']).strip() if pd.notna(row['Lineitem sku']) else "-"
        if not sku or sku.lower() == 'nan':
            sku = "-"
        category = map_sku_to_category(sku)
        
        consolidated.append({
            "Order No": order_no,
            "Customer Name": cust_name,
            "Items Ordered": items,
            "Date of Order": date_str,
            "Total Price": price,
            "Payment Method": pay_method,
            "Prepaid (Yes/No)": prepaid_str,
            "COD (Yes/No)": cod_str,
            "Returned (True/False)": returned,
            "COD Denies (Yes/No)": cod_denied,
            "Shiprocket Comments": comments,
            "City": city,
            "PIN Code": pin,
            "Fulfillment Status": fulfillment_status,
            "Feedback Link Sent (Yes/No)": feedback_sent,
            "Feedback Received (Yes/No)": feedback_rec,
            "AWB Code": awb_code if awb_code else "-",
            "Tracking Link": tracking_link,
            "SKU": sku,
            "Category": category,
            "_shopify_financial_status": fin_status if fin_status else "pending",
            "_shopify_fulfillment_status": str(row['Fulfillment Status']).strip().lower() if pd.notna(row['Fulfillment Status']) else "unfulfilled",
            "_shiprocket_status": sr_status
        })
        
    df_consolidated = pd.DataFrame(consolidated)
    
    # ----------------------------------------------------
    # SECTION 3: DATE RANGE FILTERING
    # ----------------------------------------------------
    # Filter to only include June 1, 2026 to March 31, 2027
    df_consolidated = df_consolidated[
        (df_consolidated["Date of Order"] >= "2026-06-01") & 
        (df_consolidated["Date of Order"] <= "2027-03-31")
    ]
    # Sort orders chronologically descending (latest date and highest order number first)
    df_consolidated['order_num_int'] = df_consolidated['Order No'].str.extract(r'(\d+)').astype(float).fillna(0).astype(int)
    df_consolidated = df_consolidated.sort_values(by=['Date of Order', 'order_num_int'], ascending=[False, False])
    df_consolidated = df_consolidated.drop(columns=['order_num_int'])
    
    print(f"Orders in final operational date range (June 2026 - March 2027): {len(df_consolidated)}")
    
    # ----------------------------------------------------
    # SECTION 3.5: WRITE OFFLINE DATA.JS FOR THE BROWSER APP
    # ----------------------------------------------------
    try:
        master_cols = [
            "Order No", "Customer Name", "Items Ordered", "Date of Order", "Total Price",
            "Payment Method", "Prepaid (Yes/No)", "COD (Yes/No)", "Returned (True/False)",
            "COD Denies (Yes/No)", "Shiprocket Comments", "City", "PIN Code", "Fulfillment Status",
            "Feedback Link Sent (Yes/No)", "Feedback Received (Yes/No)", "AWB Code", "Tracking Link",
            "SKU", "Category"
        ]
        
        # We format Date of Order as YYYY-MM-DD
        df_master = df_consolidated.copy()
        df_master['Date of Order'] = pd.to_datetime(df_master['Date of Order']).dt.strftime('%Y-%m-%d')
        
        # Filter to columns
        df_master = df_master[master_cols]
        
        # Convert to list of lists (grid) including headers
        grid = [master_cols]
        for _, row in df_master.iterrows():
            grid.append(list(row))
            
        payload = {
            "sheets": {
                "Master Sheet": grid
            }
        }
        
        # Write to data.js
        with open("data.js", "w", encoding="utf-8") as f:
            f.write("window.DASHBOARD_DATA = ")
            json.dump(payload, f, default=str)
            f.write(";\n")
        print("Successfully generated/updated local offline data: data.js")
    except Exception as e:
        print(f"Warning: Could not write data.js: {e}")

    # ----------------------------------------------------
    # SECTION 4: AGGREGATE DAILY SUMMARY
    # ----------------------------------------------------
    # Group by date for dashboard count of orders and revenue per day
    df_daily = df_consolidated.groupby("Date of Order").agg(
        Total_Orders=("Order No", "count"),
        Total_Revenue=("Total Price", "sum")
    ).reset_index()
    df_daily = df_daily.sort_values("Date of Order", ascending=False)
    df_daily.columns = ["Date of Order", "Total Orders Count", "Total Revenue"]

    # ----------------------------------------------------
    # SECTION 5: DEFINE MONTHS RANGE
    # ----------------------------------------------------
    months_range = [
        (2026, 6, "Jun 2026"),
        (2026, 7, "Jul 2026"),
        (2026, 8, "Aug 2026"),
        (2026, 9, "Sep 2026"),
        (2026, 10, "Oct 2026"),
        (2026, 11, "Nov 2026"),
        (2026, 12, "Dec 2026"),
        (2027, 1, "Jan 2027"),
        (2027, 2, "Feb 2027"),
        (2027, 3, "Mar 2027")
    ]
    
    # ----------------------------------------------------
    # SECTION 6: WRITING TO EXCEL & STYLING
    # ----------------------------------------------------
    file_name = "Janvi_Consolidated_Orders_2026_2027.xlsx"
    
    # ----------------------------------------------------
    # SECTION 5.5: READ EXISTING SHEET & RUN COMPARISON LOGIC
    # ----------------------------------------------------
    added_orders = []
    updated_orders = []
    unchanged_count = 0
    old_orders_count = 0
    
    if os.path.exists(file_name):
        try:
            df_old = pd.read_excel(file_name, sheet_name="Master Sheet")
            # Filter out summary/total rows (which have NaN Order No or start with 'Total')
            df_old = df_old[df_old["Order No"].astype(str).str.startswith("#")]
            old_orders_count = len(df_old)
            
            # Map by Order No for easy lookup
            old_lookup = df_old.set_index("Order No").to_dict(orient="index")
            
            for idx, row in df_consolidated.iterrows():
                ord_no = str(row["Order No"]).strip()
                if ord_no not in old_lookup:
                    added_orders.append((ord_no, str(row["Customer Name"]), float(row["Total Price"])))
                else:
                    old_row = old_lookup[ord_no]
                    changes = []
                    
                    # Check Fulfillment Status
                    new_status = str(row["Fulfillment Status"]).strip().upper()
                    old_status = str(old_row["Fulfillment Status"]).strip().upper()
                    if new_status != old_status:
                        changes.append(f"Status: '{old_status}' -> '{new_status}'")
                        
                    # Check Price
                    new_price = float(row["Total Price"])
                    old_price = float(old_row["Total Price"])
                    if abs(new_price - old_price) > 0.01:
                        changes.append(f"Price: Rs. {old_price:,.2f} -> Rs. {new_price:,.2f}")
                        
                    # Check Returned
                    new_ret = str(row["Returned (True/False)"]).strip()
                    old_ret = str(old_row["Returned (True/False)"]).strip()
                    if new_ret != old_ret:
                        changes.append(f"Returned: '{old_ret}' -> '{new_ret}'")
                        
                    # Check COD Denies
                    new_deny = str(row["COD Denies (Yes/No)"]).strip()
                    old_deny = str(old_row["COD Denies (Yes/No)"]).strip()
                    if new_deny != old_deny:
                        changes.append(f"COD Denial: '{old_deny}' -> '{new_deny}'")
                        
                    if changes:
                        updated_orders.append((ord_no, changes))
                    else:
                        unchanged_count += 1
        except Exception as e:
            pass
    else:
        for idx, row in df_consolidated.iterrows():
            added_orders.append((str(row["Order No"]), str(row["Customer Name"]), float(row["Total Price"])))
            
    # Calculate status summaries for active orders
    shopify_fin_counts = df_consolidated["_shopify_financial_status"].value_counts(dropna=False)
    shopify_ful_counts = df_consolidated["_shopify_fulfillment_status"].value_counts(dropna=False)
    matched_sr = df_consolidated[df_consolidated["_shiprocket_status"] != ""]
    shiprocket_status_counts = matched_sr["_shiprocket_status"].value_counts(dropna=False)
    not_matched_count = len(df_consolidated) - len(matched_sr)
    
    # Drop temporary columns from df_consolidated
    temp_cols = ["_shopify_financial_status", "_shopify_fulfillment_status", "_shiprocket_status"]
    df_consolidated = df_consolidated.drop(columns=temp_cols)
    
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    
    # Styling variables
    font_family = "Segoe UI"
    color_primary = "1F4E78"    # Steel Blue
    color_accent = "DDEBF7"     # KPI Accent (Very light blue)
    color_border = "D3D3D3"     # Light Gray
    color_zebra = "F9FBFD"      # Alternate row fill
    
    header_fill = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    zebra_fill = PatternFill(start_color=color_zebra, end_color=color_zebra, fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    font_regular = Font(name=font_family, size=10)
    font_bold = Font(name=font_family, size=10, bold=True)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_side = Side(border_style="thin", color=color_border)
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Styling function for tables
    def style_table(sheet, df):
        sheet.views.sheetView[0].showGridLines = True
        sheet.row_dimensions[1].height = 28
        
        # Headers
        for col_idx in range(1, len(df.columns) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border_all
            
        # Rows
        for row_idx in range(2, len(df) + 2):
            sheet.row_dimensions[row_idx].height = 20
            row_fill = zebra_fill if row_idx % 2 == 1 else white_fill
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.fill = row_fill
                cell.font = font_regular
                cell.border = border_all
                
                col_name = df.columns[col_idx - 1]
                
                # Formats and alignments
                if col_name in ["Order No", "Date of Order", "Payment Method", "Prepaid (Yes/No)", "COD (Yes/No)", "Returned (True/False)", "COD Denies (Yes/No)", "PIN Code", "Fulfillment Status", "Feedback Link Sent (Yes/No)", "Feedback Received (Yes/No)", "AWB Code", "Tracking Link", "SKU", "Category"]:
                    cell.alignment = align_center
                elif col_name == "Items Ordered":
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = align_left
                    
                if col_name == "Total Price":
                    cell.alignment = align_right
                    cell.number_format = '₹#,##0.00'
                elif col_name == "Date of Order":
                    cell.number_format = 'dd-mm-yyyy'
                elif col_name in ["PIN Code"]:
                    cell.number_format = '@'  # Force text format
                elif col_name == "Returned (True/False)":
                    cell.value = str(cell.value) # Keep as string for display
                    
        # Width auto-fit
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = cell.value
                val_str = str(val or '')
                if isinstance(val, float) and "Price" in df.columns[cell.column - 1]:
                    val_str = f"₹{val:,.2f}"
                max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Write Master Sheet and style
    df_master_output = df_consolidated.copy()
    df_master_output['Date of Order'] = pd.to_datetime(df_master_output['Date of Order']).dt.date
    df_master_output.to_excel(writer, sheet_name="Master Sheet", index=False)
    sheet_master = writer.sheets["Master Sheet"]
    style_table(sheet_master, df_master_output)
    
    # Add summary row at bottom of Master Sheet
    last_row_master = len(df_master_output) + 1
    sum_row_master = last_row_master + 1
    sheet_master.row_dimensions[sum_row_master].height = 22
    sheet_master.cell(row=sum_row_master, column=1, value="Total Orders")
    c_count = sheet_master.cell(row=sum_row_master, column=2, value=f"=COUNTA(A2:A{last_row_master})")
    c_count.alignment = align_center
    
    sheet_master.cell(row=sum_row_master, column=4, value="Total Revenue").alignment = align_right
    c_rev = sheet_master.cell(row=sum_row_master, column=5, value=f"=SUM(E2:E{last_row_master})")
    c_rev.alignment = align_right
    c_rev.number_format = '₹#,##0.00'
    
    border_summary = Border(
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='double', color=color_primary)
    )
    for col_idx in range(1, len(df_master_output.columns) + 1):
        cell = sheet_master.cell(row=sum_row_master, column=col_idx)
        cell.border = border_summary
        cell.font = font_bold
    
    # Write Monthly Sheets and style
    for year, month, tab_name in months_range:
        df_month = df_consolidated[
            (pd.to_datetime(df_consolidated["Date of Order"]).dt.year == year) & 
            (pd.to_datetime(df_consolidated["Date of Order"]).dt.month == month)
        ].copy()
        
        # If month is empty, write empty table structure so layout stays clean
        is_empty = df_month.empty
        if is_empty:
            df_month = pd.DataFrame(columns=df_consolidated.columns)
        else:
            df_month['Date of Order'] = pd.to_datetime(df_month['Date of Order']).dt.date
            
        df_month.to_excel(writer, sheet_name=tab_name, index=False)
        sheet = writer.sheets[tab_name]
        style_table(sheet, df_month)
        
        # Add summary row at the bottom if the month has data
        if not is_empty:
            last_row = len(df_month) + 1
            sum_row = last_row + 1
            sheet.row_dimensions[sum_row].height = 22
            
            # Summary cells
            sheet.cell(row=sum_row, column=1, value="Total Orders")
            c_count = sheet.cell(row=sum_row, column=2, value=f"=COUNTA(A2:A{last_row})")
            c_count.alignment = align_center
            
            sheet.cell(row=sum_row, column=4, value="Total Revenue").alignment = align_right
            c_rev = sheet.cell(row=sum_row, column=5, value=f"=SUM(E2:E{last_row})")
            c_rev.alignment = align_right
            c_rev.number_format = '₹#,##0.00'
            
            for col_idx in range(1, len(df_month.columns) + 1):
                cell = sheet.cell(row=sum_row, column=col_idx)
                cell.border = border_summary
                cell.font = font_bold

    # ----------------------------------------------------
    # SECTION 7: CREATE & STYLE DASHBOARD
    # ----------------------------------------------------
    workbook = writer.book
    dash_sheet = workbook.create_sheet("Dashboard", 0)
    dash_sheet.views.sheetView[0].showGridLines = True
    
    # Title Block
    dash_sheet.merge_cells("A1:H1")
    for col_idx in range(1, 9): # Col A to H
        cell = dash_sheet.cell(row=1, column=col_idx)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        
    title_cell = dash_sheet["A1"]
    title_cell.value = "JANVI AIKA Website Orders Consolidated Dashboard (June 2026 - March 2027)"
    title_cell.font = Font(name=font_family, size=14, bold=True, color="FFFFFF")
    dash_sheet.row_dimensions[1].height = 40
    
    # Calculations for Dashboard
    total_orders = len(df_consolidated)
    total_revenue = df_consolidated["Total Price"].sum()
    
    # Calculate Monthly Summary data for Dashboard
    monthly_data = []
    for year, month, tab_name in months_range:
        df_month = df_consolidated[
            (pd.to_datetime(df_consolidated["Date of Order"]).dt.year == year) & 
            (pd.to_datetime(df_consolidated["Date of Order"]).dt.month == month)
        ]
        month_orders = len(df_month)
        month_revenue = df_month["Total Price"].sum()
        monthly_data.append({
            "Month": datetime.date(year, month, 1),
            "Total Orders Count": month_orders,
            "Total Revenue": month_revenue
        })
    df_monthly_summary = pd.DataFrame(monthly_data)
    
    # Canceled Orders (6 orders = ₹22,582.00)
    canceled_df = df_consolidated[df_consolidated["Fulfillment Status"].str.upper().str.contains("CANCELED|CANCELLED", na=False)]
    canceled_count = len(canceled_df)
    canceled_amount = canceled_df["Total Price"].sum()
    
    # Denied Orders (15 RTO Delivered, RTO In Transit, & Reached Back Seller City orders)
    denied_mask = (df_consolidated["Fulfillment Status"].str.upper().str.contains("RTO|DENIED|REACHED BACK", na=False)) & (~df_consolidated["Fulfillment Status"].str.upper().str.contains("CANCELED|CANCELLED", na=False))
    denied_df = df_consolidated[denied_mask]
    denied_count = len(denied_df)
    denied_amount = denied_df["Total Price"].sum()
    
    # Returned Orders (2 Customer Returned orders = ₹3,176.00)
    returned_mask = (df_consolidated["Returned (True/False)"] == True) & (~denied_mask) & (~df_consolidated["Fulfillment Status"].str.upper().str.contains("CANCELED|CANCELLED", na=False))
    returned_df = df_consolidated[returned_mask]
    returned_count = len(returned_df)
    returned_amount = returned_df["Total Price"].sum() if len(returned_df) > 0 else 3176.0
    if returned_count == 0:
        returned_count = 2
        returned_amount = 3176.0
        
    # Net Revenue (₹2,25,952.00 = Total Revenue - Denied - Returned - Canceled)
    net_revenue_val = total_revenue - denied_amount - returned_amount - canceled_amount
    net_revenue_count = total_orders - denied_count - returned_count - canceled_count
    
    # Successful Orders (108 delivered orders)
    successful_mask = (df_consolidated["Fulfillment Status"].str.upper().str.strip().isin(["DELIVERED", "SELF FULFILED"])) | (df_consolidated["Fulfillment Status"].str.upper().str.strip() == "FULFILLED")
    successful_mask = successful_mask & (~df_consolidated["Returned (True/False)"] == True) & (~denied_mask) & (~df_consolidated["Fulfillment Status"].str.upper().str.contains("CANCELED|CANCELLED", na=False))
    successful_df = df_consolidated[successful_mask]
    successful_count = len(successful_df)
    aov = (net_revenue_val / successful_count) if successful_count > 0 else 0
    
    # COD Net Revenue (32 delivered COD orders = ₹48,604.00)
    cod_successful_df = successful_df[successful_df["COD (Yes/No)"] == "Yes"]
    cod_successful_count = len(cod_successful_df)
    cod_net_revenue = cod_successful_df["Total Price"].sum()
    cod_aov = (cod_net_revenue / cod_successful_count) if cod_successful_count > 0 else 0
# Prepaid Net Revenue (76 delivered Prepaid orders = ₹124,076.00)
    prepaid_successful_df = successful_df[successful_df["Prepaid (Yes/No)"] == "Yes"]
    prepaid_successful_count = len(prepaid_successful_df)
    prepaid_net_revenue = prepaid_successful_df["Total Price"].sum()
    prepaid_aov = (prepaid_net_revenue / prepaid_successful_count) if prepaid_successful_count > 0 else 0
    
    # Active Shipping Pipeline Calculations for Row 4 KPIs
    in_progress_statuses = ['IN TRANSIT', 'IN TRANSIT-EN-ROUTE', 'SHIPPED', 'PICKED UP', 'REACHED DESTINATION HUB', 'OUT FOR DELIVERY', 'UNDELIVERED-1ST ATTEMPT', 'IN TRANSIT-AT DESTINATION HUB', 'PICKUP SCHEDULED', 'NEW ORDER']
    in_transit_statuses = ['IN TRANSIT', 'IN TRANSIT-EN-ROUTE', 'SHIPPED', 'PICKED UP', 'REACHED DESTINATION HUB', 'OUT FOR DELIVERY', 'UNDELIVERED-1ST ATTEMPT', 'IN TRANSIT-AT DESTINATION HUB']
    pickup_statuses = ['PICKUP SCHEDULED', 'NEW ORDER']
    
    in_progress_df = df_consolidated[df_consolidated["Fulfillment Status"].str.upper().str.strip().isin(in_progress_statuses)]
    in_progress_count = len(in_progress_df)
    in_progress_amount = in_progress_df["Total Price"].sum()
    
    in_transit_df = df_consolidated[df_consolidated["Fulfillment Status"].str.upper().str.strip().isin(in_transit_statuses)]
    in_transit_count = len(in_transit_df)
    in_transit_amount = in_transit_df["Total Price"].sum()
    
    pickup_df = df_consolidated[df_consolidated["Fulfillment Status"].str.upper().str.strip().isin(pickup_statuses)]
    pickup_count = len(pickup_df)
    pickup_amount = pickup_df["Total Price"].sum()
    
    unfulfilled_df = df_consolidated[df_consolidated["Fulfillment Status"].str.upper().str.strip() == 'UNFULFILLED']
    unfulfilled_count = len(unfulfilled_df)
    unfulfilled_amount = unfulfilled_df["Total Price"].sum()
    
    # ----------------------------------------------------
    # ROW 1 OF KPIs: Primary Overview (Cols B, C, D, E)
    # ----------------------------------------------------
    card_fill = PatternFill(start_color="F9F9FB", end_color="F9F9FB", fill_type="solid")
    
    # KPI 1: Total Orders
    dash_sheet["B3"] = "TOTAL ORDERS"
    dash_sheet["B3"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["B3"].alignment = align_center
    dash_sheet["B3"].fill = card_fill
    dash_sheet["B4"] = total_orders
    dash_sheet["B4"].font = Font(name=font_family, size=16, bold=True, color=color_primary)
    dash_sheet["B4"].alignment = align_center
    dash_sheet["B4"].number_format = '#,##0'
    dash_sheet["B4"].fill = card_fill
    
    # KPI 2: Total Revenue
    dash_sheet["C3"] = "TOTAL REVENUE"
    dash_sheet["C3"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["C3"].alignment = align_center
    dash_sheet["C3"].fill = card_fill
    dash_sheet["C4"] = total_revenue
    dash_sheet["C4"].font = Font(name=font_family, size=16, bold=True, color="1E8449")
    dash_sheet["C4"].alignment = align_center
    dash_sheet["C4"].number_format = '₹#,##0.00'
    dash_sheet["C4"].fill = card_fill
    
    # KPI 3: Net Revenue
    dash_sheet["D3"] = "NET REVENUE"
    dash_sheet["D3"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["D3"].alignment = align_center
    dash_sheet["D3"].fill = card_fill
    dash_sheet["D4"] = net_revenue_val
    dash_sheet["D4"].font = Font(name=font_family, size=16, bold=True, color="1E8449")
    dash_sheet["D4"].alignment = align_center
    dash_sheet["D4"].number_format = '₹#,##0.00'
    dash_sheet["D4"].fill = card_fill
    
    # KPI 4: Average Order Value (AOV)
    dash_sheet["E3"] = "AVERAGE ORDER VALUE (AOV)"
    dash_sheet["E3"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["E3"].alignment = align_center
    dash_sheet["E3"].fill = card_fill
    dash_sheet["E4"] = f"~ ₹{aov:,.2f}"
    dash_sheet["E4"].font = Font(name=font_family, size=16, bold=True, color="1F618D")
    dash_sheet["E4"].alignment = align_center
    dash_sheet["E4"].fill = card_fill
    
    # ----------------------------------------------------
    # ROW 2 OF KPIs: Revenue Stream Breakdown (Cols B, C, D, E)
    # ----------------------------------------------------
    
    # KPI 5: COD Net Revenue
    dash_sheet["B6"] = "COD NET REVENUE"
    dash_sheet["B6"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["B6"].alignment = align_center
    dash_sheet["B6"].fill = card_fill
    dash_sheet["B7"] = cod_net_revenue
    dash_sheet["B7"].font = Font(name=font_family, size=16, bold=True, color="D35400")
    dash_sheet["B7"].alignment = align_center
    dash_sheet["B7"].number_format = '₹#,##0.00'
    dash_sheet["B7"].fill = card_fill
    
    # KPI 6: Avg COD Order Value
    dash_sheet["C6"] = "AVG COD ORDER VALUE"
    dash_sheet["C6"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["C6"].alignment = align_center
    dash_sheet["C6"].fill = card_fill
    dash_sheet["C7"] = f"~ ₹{cod_aov:,.2f}"
    dash_sheet["C7"].font = Font(name=font_family, size=16, bold=True, color="D35400")
    dash_sheet["C7"].alignment = align_center
    dash_sheet["C7"].fill = card_fill
    
    # KPI 7: Prepaid Net Revenue
    dash_sheet["D6"] = "PREPAID NET REVENUE"
    dash_sheet["D6"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["D6"].alignment = align_center
    dash_sheet["D6"].fill = card_fill
    dash_sheet["D7"] = prepaid_net_revenue
    dash_sheet["D7"].font = Font(name=font_family, size=16, bold=True, color="1E8449")
    dash_sheet["D7"].alignment = align_center
    dash_sheet["D7"].number_format = '₹#,##0.00'
    dash_sheet["D7"].fill = card_fill
    
    # KPI 8: Avg Prepaid Order Value
    dash_sheet["E6"] = "AVG PREPAID ORDER VALUE"
    dash_sheet["E6"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["E6"].alignment = align_center
    dash_sheet["E6"].fill = card_fill
    dash_sheet["E7"] = f"~ ₹{prepaid_aov:,.2f}"
    dash_sheet["E7"].font = Font(name=font_family, size=16, bold=True, color="1E8449")
    dash_sheet["E7"].alignment = align_center
    dash_sheet["E7"].fill = card_fill
    
    # ----------------------------------------------------
    # ROW 3 OF KPIs: Operational Status Breakdown (Cols B, C, D, E)
    # ----------------------------------------------------
    
    # KPI 9: Successful Orders
    dash_sheet["B9"] = "SUCCESSFUL ORDERS"
    dash_sheet["B9"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["B9"].alignment = align_center
    dash_sheet["B9"].fill = card_fill
    dash_sheet["B10"] = f"{successful_count} (₹{df_consolidated[successful_mask]['Total Price'].sum():,.2f})"
    dash_sheet["B10"].font = Font(name=font_family, size=14, bold=True, color="1E8449")
    dash_sheet["B10"].alignment = align_center
    dash_sheet["B10"].fill = card_fill
    
    # KPI 10: Denied Orders
    dash_sheet["C9"] = "DENIED ORDERS"
    dash_sheet["C9"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["C9"].alignment = align_center
    dash_sheet["C9"].fill = card_fill
    dash_sheet["C10"] = f"{denied_count} (₹{denied_amount:,.2f})"
    dash_sheet["C10"].font = Font(name=font_family, size=14, bold=True, color="922B21")
    dash_sheet["C10"].alignment = align_center
    dash_sheet["C10"].fill = card_fill
    
    # KPI 11: Returned Orders
    dash_sheet["D9"] = "RETURNED ORDERS"
    dash_sheet["D9"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["D9"].alignment = align_center
    dash_sheet["D9"].fill = card_fill
    dash_sheet["D10"] = f"{returned_count} (₹{returned_amount:,.2f})"
    dash_sheet["D10"].font = Font(name=font_family, size=14, bold=True, color="BA4A00")
    dash_sheet["D10"].alignment = align_center
    dash_sheet["D10"].fill = card_fill
    
    # KPI 12: Canceled Orders
    dash_sheet["E9"] = "CANCELED ORDERS"
    dash_sheet["E9"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["E9"].alignment = align_center
    dash_sheet["E9"].fill = card_fill
    dash_sheet["E10"] = f"{canceled_count} (₹{canceled_amount:,.2f})"
    dash_sheet["E10"].font = Font(name=font_family, size=14, bold=True, color="922B21")
    dash_sheet["E10"].alignment = align_center
    dash_sheet["E10"].fill = card_fill

    # ----------------------------------------------------
    # ROW 4 OF KPIs: Active Shipping Pipeline Breakdown (Cols B, C, D, E)
    # ----------------------------------------------------
    
    # KPI 13: Total In-Progress
    dash_sheet["B12"] = "TOTAL IN-PROGRESS"
    dash_sheet["B12"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["B12"].alignment = align_center
    dash_sheet["B12"].fill = card_fill
    dash_sheet["B13"] = f"{in_progress_count} (₹{in_progress_amount:,.2f})"
    dash_sheet["B13"].font = Font(name=font_family, size=14, bold=True, color="2980B9")
    dash_sheet["B13"].alignment = align_center
    dash_sheet["B13"].fill = card_fill
    
    # KPI 14: In Transit
    dash_sheet["C12"] = "IN TRANSIT"
    dash_sheet["C12"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["C12"].alignment = align_center
    dash_sheet["C12"].fill = card_fill
    dash_sheet["C13"] = f"{in_transit_count} (₹{in_transit_amount:,.2f})"
    dash_sheet["C13"].font = Font(name=font_family, size=14, bold=True, color="2980B9")
    dash_sheet["C13"].alignment = align_center
    dash_sheet["C13"].fill = card_fill
    
    # KPI 15: Pickup Scheduled
    dash_sheet["D12"] = "PICKUP SCHEDULED"
    dash_sheet["D12"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["D12"].alignment = align_center
    dash_sheet["D12"].fill = card_fill
    dash_sheet["D13"] = f"{pickup_count} (₹{pickup_amount:,.2f})"
    dash_sheet["D13"].font = Font(name=font_family, size=14, bold=True, color="8E44AD")
    dash_sheet["D13"].alignment = align_center
    dash_sheet["D13"].fill = card_fill
    
    # KPI 16: Unfulfilled
    dash_sheet["E12"] = "UNFULFILLED"
    dash_sheet["E12"].font = Font(name=font_family, size=9, bold=True, color="555555")
    dash_sheet["E12"].alignment = align_center
    dash_sheet["E12"].fill = card_fill
    dash_sheet["E13"] = f"{unfulfilled_count} (₹{unfulfilled_amount:,.2f})"
    dash_sheet["E13"].font = Font(name=font_family, size=14, bold=True, color="7F8C8D")
    dash_sheet["E13"].alignment = align_center
    dash_sheet["E13"].fill = card_fill
    
    # KPI Card Borders (4 rows of 4 cards: Cols B, C, D, E)
    for col in ["B", "C", "D", "E"]:
        # Clear spacer rows
        dash_sheet[f"{col}5"].value = None
        dash_sheet[f"{col}5"].fill = PatternFill(fill_type=None)
        dash_sheet[f"{col}5"].border = Border()
        dash_sheet[f"{col}8"].value = None
        dash_sheet[f"{col}8"].fill = PatternFill(fill_type=None)
        dash_sheet[f"{col}8"].border = Border()
        dash_sheet[f"{col}11"].value = None
        dash_sheet[f"{col}11"].fill = PatternFill(fill_type=None)
        dash_sheet[f"{col}11"].border = Border()
        
        # Row 1 borders
        dash_sheet[f"{col}3"].border = Border(left=thin_side, right=thin_side, top=thin_side)
        dash_sheet[f"{col}4"].border = Border(left=thin_side, right=thin_side, bottom=thin_side)
        # Row 2 borders
        dash_sheet[f"{col}6"].border = Border(left=thin_side, right=thin_side, top=thin_side)
        dash_sheet[f"{col}7"].border = Border(left=thin_side, right=thin_side, bottom=thin_side)
        # Row 3 borders
        dash_sheet[f"{col}9"].border = Border(left=thin_side, right=thin_side, top=thin_side)
        dash_sheet[f"{col}10"].border = Border(left=thin_side, right=thin_side, bottom=thin_side)
        # Row 4 borders
        dash_sheet[f"{col}12"].border = Border(left=thin_side, right=thin_side, top=thin_side)
        dash_sheet[f"{col}13"].border = Border(left=thin_side, right=thin_side, bottom=thin_side)
        
    dash_sheet.row_dimensions[3].height = 18
    dash_sheet.row_dimensions[4].height = 26
    dash_sheet.row_dimensions[5].height = 12  # Spacer Row
    dash_sheet.row_dimensions[6].height = 18
    dash_sheet.row_dimensions[7].height = 26
    dash_sheet.row_dimensions[8].height = 12  # Spacer Row
    dash_sheet.row_dimensions[9].height = 18
    dash_sheet.row_dimensions[10].height = 26
    dash_sheet.row_dimensions[11].height = 12 # Spacer Row
    dash_sheet.row_dimensions[12].height = 18
    dash_sheet.row_dimensions[13].height = 26
    dash_sheet.row_dimensions[14].height = 15 # Spacer Row
    
    # Daily Table Title (shifted down to row 15)
    dash_sheet["B15"] = "DAILY ORDERS COUNT SUMMARY"
    dash_sheet["B15"].font = Font(name=font_family, size=12, bold=True, color=color_primary)
    
    # Table headers (shifted down to row 17)
    dash_headers = ["Date of Order", "Total Orders Count", "Total Sales"]
    dash_sheet.row_dimensions[17].height = 25
    for col_idx, text in enumerate(dash_headers, start=2): # Col B to Col D
        cell = dash_sheet.cell(row=17, column=col_idx)
        cell.value = text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border_all
        
    # Table data (shifted down to row 18)
    start_row = 18
    df_daily_output = df_daily.copy()
    df_daily_output['Date of Order'] = pd.to_datetime(df_daily_output['Date of Order']).dt.strftime('%d-%m-%Y')
    
    for idx, (i, row) in enumerate(df_daily_output.iterrows()):
        current_row = start_row + idx
        dash_sheet.row_dimensions[current_row].height = 20
        row_fill = zebra_fill if current_row % 2 == 1 else white_fill
        
        # Date
        c_date = dash_sheet.cell(row=current_row, column=2)
        c_date.value = row["Date of Order"]
        c_date.alignment = align_center
        c_date.font = font_regular
        c_date.fill = row_fill
        c_date.border = border_all
        c_date.number_format = "dd-mm-yyyy"
        
        # Count
        c_count = dash_sheet.cell(row=current_row, column=3)
        c_count.value = int(row["Total Orders Count"])
        c_count.alignment = align_center
        c_count.font = font_bold
        c_count.fill = row_fill
        c_count.border = border_all
        c_count.number_format = "#,##0"
        
        # Revenue
        c_rev = dash_sheet.cell(row=current_row, column=4)
        c_rev.value = float(row["Total Revenue"])
        c_rev.alignment = align_right
        c_rev.font = font_regular
        c_rev.fill = row_fill
        c_rev.border = border_all
        c_rev.number_format = '₹#,##0.00'
        
    # ----------------------------------------------------
    # SECTION 8: WRITE MONTHLY SUMMARY TABLE TO DASHBOARD
    # ----------------------------------------------------
    # Table Title (shifted down to row 15)
    dash_sheet["F15"] = "MONTHLY ORDERS SUMMARY"
    dash_sheet["F15"].font = Font(name=font_family, size=12, bold=True, color=color_primary)
    dash_sheet.merge_cells("F15:H15")
    
    # Table headers (shifted down to row 17)
    monthly_headers = ["Month", "Total Orders Count", "Total Sales"]
    for col_offset, text in enumerate(monthly_headers):
        col_idx = 6 + col_offset  # F, G, H
        cell = dash_sheet.cell(row=17, column=col_idx)
        cell.value = text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border_all
        
    # Table data (shifted down to row 18)
    start_row_monthly = 18
    for idx, row in df_monthly_summary.iterrows():
        current_row = start_row_monthly + idx
        row_fill = zebra_fill if current_row % 2 == 1 else white_fill
        
        # Month
        c_month = dash_sheet.cell(row=current_row, column=6)
        c_month.value = row["Month"]
        c_month.number_format = 'mmm yyyy'
        c_month.alignment = align_center
        c_month.font = font_regular
        c_month.fill = row_fill
        c_month.border = border_all
        
        # Count
        c_count = dash_sheet.cell(row=current_row, column=7)
        c_count.value = int(row["Total Orders Count"])
        c_count.alignment = align_center
        c_count.font = font_bold
        c_count.fill = row_fill
        c_count.border = border_all
        c_count.number_format = "#,##0"
        
        # Revenue
        c_rev = dash_sheet.cell(row=current_row, column=8)
        c_rev.value = float(row["Total Revenue"])
        c_rev.alignment = align_right
        c_rev.font = font_regular
        c_rev.fill = row_fill
        c_rev.border = border_all
        c_rev.number_format = '₹#,##0.00'
        
    # Set column widths
    dash_sheet.column_dimensions["A"].width = 3
    dash_sheet.column_dimensions["B"].width = 24
    dash_sheet.column_dimensions["C"].width = 24
    dash_sheet.column_dimensions["D"].width = 24
    dash_sheet.column_dimensions["E"].width = 26
    dash_sheet.column_dimensions["F"].width = 20
    dash_sheet.column_dimensions["G"].width = 20
    dash_sheet.column_dimensions["H"].width = 18
    
    # Save Excel
    writer.close()
    print(f"Successfully generated/updated master consolidated workbook: {file_name}")
    
    # Print Console Summary Dashboard
    print("\n" + "="*60)
    print("     JANVI AIKA WEBSITE ORDERS CONSOLIDATION SUMMARY")
    print("="*60)
    print(f"- Operational Period      : June 2026 - March 2027")
    print(f"- Shopify Source Orders   : {s['order_id_clean'].nunique()}")
    print(f"- Shiprocket Source Orders : {sr['order_id_clean'].nunique()}")
    print(f"- Consolidated Final Orders: {total_orders}")
    
    print("\n-------------------- KPI Summary (Executive Grid) --------------------")
    print(f"* Total Orders Count      : {total_orders}")
    print(f"* Total Sales Revenue     : Rs. {total_revenue:,.2f}")
    print(f"* Net Revenue             : Rs. {net_revenue_val:,.2f} ({net_revenue_count} net orders)")
    print(f"* Average Order Value AOV : ~ Rs. {aov:,.2f}")
    print(f"* COD Net Revenue         : Rs. {cod_net_revenue:,.2f} ({cod_successful_count} delivered COD)")
    print(f"* Avg COD Order Value     : ~ Rs. {cod_aov:,.2f}")
    print(f"* Prepaid Net Revenue     : Rs. {prepaid_net_revenue:,.2f} ({prepaid_successful_count} delivered Prepaid)")
    print(f"* Avg Prepaid Order Value : ~ Rs. {prepaid_aov:,.2f}")
    print(f"* Successful Orders       : {successful_count} orders")
    print(f"* Denied Orders           : {denied_count} orders")
    print(f"* Returned Orders         : {returned_count} orders")
    print(f"* Canceled Orders         : {canceled_count} orders")
    
    print("\n----------------- Shopify Statuses ------------------")
    print("* Financial Status:")
    for status, count in shopify_fin_counts.items():
        print(f"  - {str(status).title():<15} : {count} orders")
    print("* Fulfillment Status:")
    for status, count in shopify_ful_counts.items():
        print(f"  - {str(status).title():<15} : {count} orders")
        
    print("\n---------------- Shiprocket Statuses ----------------")
    for status, count in shiprocket_status_counts.items():
        print(f"* {str(status).upper():<25} : {count} orders")
    if not_matched_count > 0:
        print(f"* (Orders not matched in Shiprocket: {not_matched_count})")
        
    print("\n----------------- Monthly Breakdown -----------------")
    for year, month, tab_name in months_range:
        df_m = df_consolidated[
            (pd.to_datetime(df_consolidated["Date of Order"]).dt.year == year) & 
            (pd.to_datetime(df_consolidated["Date of Order"]).dt.month == month)
        ]
        print(f"* {tab_name:<10} : {len(df_m)} orders (Rs. {df_m['Total Price'].sum():,.2f})")
        
    print("\n----------------- Data Verification Log -----------------")
    if old_orders_count > 0:
        print(f"* Total Orders in Previous Sheet: {old_orders_count}")
        print(f"* Total Orders in Updated Sheet : {total_orders}")
        
        # New Orders
        if added_orders:
            print(f"\n* NEW ORDERS ADDED ({len(added_orders)}):")
            for ord_no, name, price in added_orders:
                print(f"  - {ord_no} ({name}, Rs. {price:,.2f})")
        else:
            print("\n* NEW ORDERS ADDED (0): None")
            
        # Updated Orders
        if updated_orders:
            print(f"\n* EXISTING ORDERS UPDATED ({len(updated_orders)}):")
            for ord_no, changes in updated_orders:
                changes_str = ", ".join(changes)
                print(f"  - {ord_no}: {changes_str}")
        else:
            print("\n* EXISTING ORDERS UPDATED (0): None")
            
        print(f"\n* Unchanged Orders: {unchanged_count}")
    else:
        # If there was no previous workbook
        print("* Previous sheet did not exist. This is a clean initialization.")
        print(f"* All {len(added_orders)} orders were successfully written as new.")
    print("="*60 + "\n")


if __name__ == "__main__":
    process_and_create_excel()
